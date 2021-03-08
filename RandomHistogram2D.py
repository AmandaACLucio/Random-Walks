

############################################ Inputs #############################################

while True:
    try:
        macroporousSize=int(input("What is the macropore radius (R)? "))
        if not 0 < macroporousSize:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break

while True:
    try:
        microporousSize=int(input("What is the square width (W)? "))
        if macroporousSize >= microporousSize:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break

while True:
    try:
        numberMaxStepsMacropore=int(input("What is the maximum number of steps (macropore)? "))
        if not 0 < numberMaxStepsMacropore:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break

while True:
    try:
        numberMaxStepsMicropore=int(input("What is the maximum number of steps (micropore)? "))
        if not 0 < numberMaxStepsMicropore:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break

while True:
    try:
        numberWalks=int(input("How many random walks? "))
        if not 0 < numberWalks:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break

choose=0
while True:
    try:
        print("Choose an option:")
        print("1. Random walker should start in the middle")
        print("2. Random walker should start inside the macropore")
        print("3. Random walker should start anywhere (micro and macroporosity)?")
        choose=int(input("Enter your option number: "))
        if not ((choose==1) or (choose==2)or (choose==3)):
            raise ValueError(choose)
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break  


followingRule ="no"
if choose==3:
    while True:
        try:
            followingRule=input("Should microporosity random walkers difuse into the macroporosity? [yes/no] ").lower()
            if not ((followingRule=="yes") or (followingRule=="no")):
                raise ValueError(followingRule)
        except ValueError as e:
            print("Invalid value:", e)
        else:
            break


porcetageMicropore=0
if choose==3:
    while True:
        try:
            porcetageMicropore=int(input("What is the percentage (%) of random walkers in the micropore? [only number] "))
            if not ((porcetageMicropore>=0) or (porcetageMicropore<=100)):
                raise ValueError(porcetageMicropore)
        except ValueError as e:
            print("Invalid value:", e)
        else:
            break

while True:
    try:
        minimumTime2=float(input(u"What is the T2min? "))
        if not 0 < minimumTime2:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break   

while True:
    try:
        maximumTime2=int(input(u"What is the T2max? "))
        if not minimumTime2 < maximumTime2:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break    

while True:
    try:
        numberBins=int(input(u"What is the number of bins? "))
        if not 0 < numberBins:
            raise ValueError("Enter a valid value")
    except ValueError as e:
        print("Invalid value:", e)
    else:
        break   

if followingRule=="no":
    followingRule=False
else:
    followingRule=True



############################################ imports #########################################


# I changed the location to become the code more fast
import numpy 
import pylab 
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment


############################################ Class #############################################

class Environment:

    def __init__(self, solidShape, size, numberMaxSteps):

        self.size           = size
        self.solidShape     = solidShape
        self.numberMaxSteps = numberMaxSteps

    
    def inside(self, x, y):
        
        if self.solidShape =="micropore":

            widthMiddle = self.size/2
            
            if (x>widthMiddle) or (x<(-1*widthMiddle)) or (y>widthMiddle) or (y<(-1*widthMiddle)):
            
                return False
            return True

        elif self.solidShape =="macropore":

            if  numpy.sqrt((x**2)+(y**2))>=self.size:
                return False
            return True

class RandomWalk():

    def __init__(self, countStepsMacropore, countStepsMicropore, micropore, macropore, count, x, y, initialStep):
    
        self.countStepsMacropore = countStepsMacropore #total de passos no microporo
        self.countStepsMicropore = countStepsMicropore #total de passos no microporo
        self.micropore            = micropore
        self.macropore            = macropore 
        self.count                = count #total de passos
        self.x                    = x
        self.y                    = y
        self.initialStep          = initialStep

    def square(self, valueX, valueY):
        if(self.macropore.inside(valueX, valueY)):
            return 0
        if(self.micropore.inside(valueX, valueY)):
            return 1
        return -1

    def createStepInicial(self, choose):
        if choose==1:
            self.initialStep = [0, 0]
        elif choose==2:
            radius = self.macropore.size
            self.initialStep = [random.randint(-1*radius, radius),random.randint(-1*radius, radius)]
            while(self.square(self.initialStep[0],self.initialStep[1])==1):
                self.initialStep = [random.randint(-1*radius, radius),random.randint(-1*radius, radius)]
        elif choose==3:
            width = int(self.micropore.size/2)
            self.initialStep = [random.randint(-1*width, width), random.randint(-1*width, width)]
            while(self.square(self.initialStep[0],self.initialStep[1])==0):
                self.initialStep = [random.randint(-1*width, width), random.randint(-1*width, width)]

        return 

    def walkingMacropore(self):

        if(len(self.x)==0): #it never been in micropore
    
            ## x0,y0 e z0 are the initial steps
            self.x.append(self.initialStep[0])
            self.y.append(self.initialStep[1])

            ## x1, x2, ... e y1, y2,... equal to the previous value plus 1 or minus 1. 
            ## When we draw a random number> = 0 and <0, we add -1
            ## If we randomly draw a random number> = 0, we add 1

            for step in range(1, self.macropore.numberMaxSteps+1):

                randomX = random.choice([-1, 1])
                randomY = random.choice([-1, 1])

                if randomX<0:
                    valueX = self.x[step-1]-1
                else:
                    valueX = self.x[step-1]+1

                if randomY<0:
                    valueY = self.y[step-1]-1                    
                else:
                    valueY = self.y[step-1]+1

                square = self.square(valueX, valueY)

                self.x.append(valueX)
                self.y.append(valueY)
                self.countStepsMacropore+=1
                self.count+=1

                if square==1:
                    
                    self.x = self.x[1:]
                    self.y = self.y[1:]
                    return 
            
            self.x = self.x[1:len(self.x)]
            self.y = self.y[1:len(self.y)]
            return

        else: #it was in micropore

            stepsPrevious = len(self.x)

            for step in range(stepsPrevious, stepsPrevious+self.macropore.numberMaxSteps):

                randomX = random.choice([-1, 1])
                randomY = random.choice([-1, 1])

                if randomX<0:
                    valueX = self.x[step-1]-1
                else:
                    valueX = self.x[step-1]+1

                if randomY<0:
                    valueY = self.y[step-1]-1                    
                else:
                    valueY = self.y[step-1]+1

                square = self.square(valueX, valueY)

                self.x.append(valueX)
                self.y.append(valueY)
                self.countStepsMacropore+=1
                self.count+=1                
                    
                if square==1:
                    return 

            return

    def walkingMicropore(self, followingRule):
    
        ## x0 and y0 are the initial steps
        self.x.append(self.initialStep[0])
        self.y.append(self.initialStep[1])

        ## x1, x2, ... e y1, y2,... equal to the previous value plus 1 or minus 1. 
        ## When we draw a random number> = 0 and <0, we add -1
        ## If we randomly draw a random number> = 0, we add 1

        for step in range(1, self.micropore.numberMaxSteps+1):

            randomX = random.choice([-1, 1])
            randomY = random.choice([-1, 1])

            if randomX<0:
                valueX = self.x[step-1]-1
            else:
                valueX = self.x[step-1]+1

            if randomY<0:
                valueY = self.y[step-1]-1                    
            else:
                valueY = self.y[step-1]+1

            square = self.square(valueX, valueY)

            self.x.append(valueX)
            self.y.append(valueY)
            self.count+=1
            self.countStepsMicropore+=1
            
            if square==0:
                self.x = self.x[1:]
                self.y = self.y[1:]
                if followingRule==True:
                    self.walkingMacropore()
                return

            if (square==-1):
                self.x = self.x[1:]
                self.y = self.y[1:]
                return
        
        self.x = self.x[1:]
        self.y = self.y[1:]
  
    def walking(self, choose, followingRule):

        self.createStepInicial(choose)
        square = self.square(self.initialStep[0], self.initialStep[1])

        if(square==0):
            return self.walkingMacropore()
        return self.walkingMicropore(followingRule)


######################################### Write Files ###############################################

class RandomWalkWalking():

    def __init__(self, numberWalks, porcetageMicropore,microporousSize, macroporousSize, numberMaxStepsMicropore, numberMaxStepsMacropore, followingRule, numberBins, minimumTime2, maximumTime2, timeTwoList=[], frequencyList=[], fileTxt="", fileExcel="", sheetExcel="", totalWalks=1, sumCountSteps=0, lastStepCount=[], lastStepX=[], lastStepY=[]):

        self.numberWalks = numberWalks 
        self.porcetageMicropore = porcetageMicropore
        self.macropore = "" #enviroment macropore
        self.micropore = "" #enviroment micropore
        self.microporousSize = microporousSize
        self.macroporousSize = macroporousSize
        self.numberMaxStepsMicropore = numberMaxStepsMicropore
        self.numberMaxStepsMacropore = numberMaxStepsMacropore
        self.followingRule = followingRule
        self.numberBins = numberBins
        self.minimumTime2 = minimumTime2
        self.maximumTime2 = maximumTime2
        self.timeTwoList = timeTwoList #list of times
        self.frequencyList =frequencyList #list of time'frequency
        self.fileTxt = fileTxt
        self.fileExcel = fileExcel
        self.sheetExcel = sheetExcel
        self.totalWalks = totalWalks #count of walks
        self.sumCountSteps = sumCountSteps #sum of count steps
        self.lastStepCount = lastStepCount #list with lasts steps count
        self.lastStepX = lastStepX
        self.lastStepY = lastStepY

    def createEnvironment(self):
    
        self.micropore  = Environment("micropore", self.microporousSize, self.numberMaxStepsMicropore)
        self.macropore  = Environment("macropore", self.macroporousSize, self.numberMaxStepsMacropore)
        

    def createWalk(self):
        
        countStepsMacropore=0
        countStepsMicropore=0 
        count=0
        x, y, initialStep = [], [], []

        walk = RandomWalk(countStepsMacropore, countStepsMicropore, self.micropore, self.macropore, count, x, y, initialStep)

        return walk


    def walkingOneRandomWalk(self, choose):

        self.createEnvironment()
        walk = self.createWalk()

        walk.walking(choose, self.followingRule)
        
        return walk


    def walkinRandomWalkChoose(self, choose):

        walk = self.walkingOneRandomWalk(choose)

        self.fileTxt.write("Walk: "+str(self.totalWalks)+"    x: "+ str(walk.initialStep[0])+"    y: "+ str(walk.initialStep[1])+"    step initial   environment: "+ ("macropore" if walk.macropore.inside(walk.initialStep[0], walk.initialStep[1]) else "micropore")+"    square: "+ str(walk.square(walk.initialStep[0], walk.initialStep[1]))+"\n")

        self.lastStepCount.append(walk.count) #list with all total step quantities
        self.lastStepX.append(walk.x[walk.count-1]) #list with the value of x for the last step of each random walk
        self.lastStepY.append(walk.y[walk.count-1]) #list with the value of y for the last step of each random walk

        for step in range(walk.count):

            self.fileTxt.write("Walk: "+str(self.totalWalks)+"    "+"x: "+ str(walk.x[step])+"    y: "+ str(walk.y[step])+"        step: "+str(step+1)+"    environment: "+ ("macropore" if walk.macropore.inside(walk.x[step], walk.y[step]) else "micropore")+"    square: "+ str(walk.square(walk.x[step], walk.y[step]))+"\n")
        
        self.fileTxt.write("-------------------------------------------------------------------------------------------\n")
        self.totalWalks+=1
        self.sumCountSteps+=walk.count

        return walk

    def startFile(self):

        self.fileTxt = open('RandomWalk-R'+str(self.macroporousSize)+'-W'+str(self.microporousSize)+'-S'+str(self.numberMaxStepsMacropore+self.numberMaxStepsMicropore)+'-RW'+str(self.numberWalks)+'.txt', 'w')
        self.fileTxt.write("RandomWalk  -  Raio: "+str(macroporousSize)+"  -  Wiidth: "+str(self.microporousSize)+ "    -    Steps: "+str(self.numberMaxStepsMacropore+self.numberMaxStepsMicropore)+"       -       Random Walks: "+str(self.numberWalks)+"\n")
        self.fileTxt.write("===========================================================================================\n\n")

    def endFile(self):
            
        mediaSteps = self.sumCountSteps/self.numberWalks
        macroporosity = (numpy.pi*(self.macroporousSize**2))/(self.microporousSize**2)

        self.fileTxt.write("===========================================================================================\n")
        self.fileTxt.write("The average step for "+str(self.numberWalks)+" random walkers is "+str(mediaSteps))
        self.fileTxt.write("The macroporosity is: "+"{:.2%}".format(macroporosity))
        self.fileTxt.close()

        print("The average step for "+str(self.numberWalks)+" random walkers is "+str(mediaSteps))
        print("The macroporosity is: "+"{:.2%}".format(macroporosity))


    def startfileExcel(self):

        self.fileExcel = Workbook()
        self.sheetExcel = self.fileExcel.active
        self.sheetExcel.title = 'RandomWalk-Bins'+str(self.numberBins)


    def createCellfileExcel(self, paramenters):
    
        for cell in paramenters:

            self.sheetExcel.cell(row=cell[0], column=cell[1], value=cell[2])
            self.sheetExcel.cell(row=cell[0], column=cell[1]).alignment=Alignment(horizontal="center", vertical="center")


    def buildingHistogram(self):

        for bin in range(self.numberBins):
            timeTwoValue = self.minimumTime2*((self.maximumTime2/self.minimumTime2)**(bin/(self.numberBins-1)))
            self.timeTwoList.append(timeTwoValue)

        #Frequency
        self.frequencyList = numpy.zeros((self.numberBins,), dtype=int) 

        for bin in range(len(self.timeTwoList)):
            for walk in range(len(self.lastStepX)):
                if (bin==0):
                    if (self.lastStepCount[walk]>0) and (self.lastStepCount[walk]<=self.timeTwoList[bin]):
                        self.frequencyList[bin]+=1   
                if (self.lastStepCount[walk]>self.timeTwoList[bin-1]) and (self.lastStepCount[walk]<=self.timeTwoList[bin]):
                    self.frequencyList[bin]+=1   

        #graphic
        pylab.plot(self.timeTwoList, self.frequencyList, 'o')

        pylab.title('Histogram of Random Walks',fontsize=15)

        pylab.xlabel('T2 (msec)', fontsize=12)

        scaleY = []

        maxFrequencys = max(self.frequencyList)

        division = int(maxFrequencys/15)/100
        round(division+0.5, 1) 
        division = division*100

        maxEscalay = int(max(self.frequencyList)/division)*division+division
        for i in range(int(maxEscalay/division)+1):
            scaleY.append(i*division)

        pylab.xscale('log')
        pylab.yticks(scaleY)
        pylab.grid(True)
        pylab.savefig('RandomWalk-R'+str(self.macroporousSize)+'-W'+str(self.microporousSize)+'-S'+str(self.numberMaxStepsMacropore+self.numberMaxStepsMicropore)+'-RW'+str(self.numberWalks)+'.png',dpi=600) 
        pylab.show()

    def writingExcel(self):

        self.buildingHistogram()
        self.startfileExcel()

        paramenters = [[1, 1, "Walk"],[1, 2, "X"],[1, 3, "Y"],[1, 4, "Steps"], [1, 7, "T2min"], [1, 8, self.minimumTime2], [2, 7, "T2max"], [2, 8, self.maximumTime2], [3, 7, "numberBins"], [3, 8, self.numberBins], [1, 10, "I"], [1, 11, "T2"], [1, 12, "Freq."]]
        self.createCellfileExcel(paramenters)

        for walk in range(len(self.lastStepX)):

            paramenters = [[walk+2, 1, walk+1], [walk+2, 2, self.lastStepX[walk]], [walk+2, 3, self.lastStepY[walk]],[walk+2, 4, self.lastStepCount[walk]]]
            
            self.createCellfileExcel(paramenters)

        for i in range(len(self.frequencyList)):
            
            paramenters = [[i+2, 10, i+1], [i+2, 11, self.timeTwoList[i]], [i+2, 12, self.frequencyList[i]]]
            
            self.createCellfileExcel(paramenters)

        histogram=Image('RandomWalk-R'+str(self.macroporousSize)+'-W'+str(self.microporousSize)+'-S'+str(self.numberMaxStepsMacropore+self.numberMaxStepsMicropore)+'-RW'+str(self.numberWalks)+'.png')
        histogram.width = histogram.width/7
        histogram.height =histogram.height/7

        self.sheetExcel.add_image(histogram, 'N2')
        self.fileExcel.save('RandomWalk-R'+str(self.macroporousSize)+'-W'+str(self.microporousSize)+'-S'+str(self.numberMaxStepsMacropore+self.numberMaxStepsMicropore)+'-RW'+str(self.numberWalks)+'.xlsx')

    def walkingRandomWalks(self, choose):

        self.startFile()

        numberWalksMicropore = int((self.porcetageMicropore/100)*self.numberWalks)
        numberWalksMacropore = self.numberWalks-int((self.porcetageMicropore/100)*self.numberWalks)

        if(choose!=1):
            for walk in range(numberWalksMicropore):
                walk = self.walkinRandomWalkChoose(3)

            for walk in range(numberWalksMacropore):
                walk = self.walkinRandomWalkChoose(2)

            self.endFile()
            self.writingExcel()

        else:

            for walk in range(self.numberWalks):
                walk = self.walkinRandomWalkChoose(1)

            self.endFile()
            self.writingExcel()


groupRandomWalks = RandomWalkWalking(numberWalks, porcetageMicropore,microporousSize, macroporousSize,numberMaxStepsMicropore, numberMaxStepsMacropore, followingRule, numberBins, minimumTime2, maximumTime2)
groupRandomWalks.walkingRandomWalks(choose)